# -*- coding: utf-8 -*-
"""
Created on Mon Jul  3 11:58:49 2023

@author: Franz Burauer
"""

from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from Listenelemente_final import Position
from Dateiauswahl import fallunterscheidung

def excel(file):
    versuche = fallunterscheidung(file)
    values = Position(versuche[0],versuche[1])

    key1 = values[0]
    key2 = values[1]
    liste_einstellversuche = values[2]
    liste_nachweisversuche = values[3]
    gesamtliste = versuche[2]
    wb = Workbook()
    unterstriche = [2,7,10,13,16,19,22,25,28,31,34,37,40,43]
    # grab the active worksheet
    ws = wb.active

    #zuordnung der ersten Parameter in Tabelle

    ws['A1'] = "Gruppe 04 - Zusammenfassung:"
    ws["A3"] = "Zuordnender Versuchsparameter:"
    ws["A4"] = "Key-Platzhalter (Buchstaben):"
    ws["A5"] = "Positionen des Keys im Dateinamen:"
    ws["B3"] = "Versuchsteil"
    ws["B4"] = f"{liste_einstellversuche[0][key1]}{liste_einstellversuche[0][key2]}"
    ws["B5"] = f"[{key1},{key2}]"

    #Legende der Ausprägunges des Keys
    ws["A9"] = "Legende für die verschiedenen Ausprägungen dieses Keys:"
    ws["A10"] = "Ausprägung"
    ws["B10"] = "Bedeutung"
    ws["A11"] = f"{liste_einstellversuche[0][key1]}{liste_einstellversuche[0][key2]}"
    ws["B11"] = "Einstellversuch"
    ws["A12"] = f"{liste_nachweisversuche[0][key1]}{liste_nachweisversuche[0][key2]}"
    ws["B12"] = "Nachweisversuch"

    #Problem:versuche ist noch eine kalammer mehr drin,  also 2 listen, wie raus ohne viel aufwand??
    i = 0
    dateinamen = []
    while i < len(gesamtliste):
        zwischenliste = []
        n=9#n vielleicht ersetzen duch längencode von enes aus Listenelemente_final wg fehlerunabhängiger
        zwischenliste.extend([i,gesamtliste[i]])
        while n < (len(gesamtliste[i])-3):
            zwischenliste.extend([f"{gesamtliste[i][n-1]}{gesamtliste[i][n]}"])
            n+=3
        if gesamtliste[i] in liste_einstellversuche:
            zwischenliste.append("Einstellversuch")
        else:
            zwischenliste.append("Nachweisversuch")
        dateinamen.append(zwischenliste)

        i+=1

    # add column headings. NB. these must be strings
    ws.append(["Nr.","Dateiname","AA","BB","CC","DD","EE","FF","GG","HH","II","JJ","KK","LL","MM","Zugeordneter_Versuchsteil"])
    tab = Table(displayName="Table2", ref="A15:P23")

    ws.add_table(tab)
    # Save the file
    wb.save("sample.xlsx")